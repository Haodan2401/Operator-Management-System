from django.shortcuts import render, redirect
from app01 import models
from app01.utils.pagination import Pagination
from app01.utils.form import UserModelForm, PrettyAddModelForm, PrettyEditModelForm
from django.http import JsonResponse, HttpResponse
from openpyxl import load_workbook

############## Department Form ####################

def depart_list(request):
    """ department list """
    queryset = models.Department.objects.all()
    page_obj = Pagination(request, queryset, page_size=2)

    context = {
        "queryset": page_obj.page_queryset,
        "page_string": page_obj.html,
    }
    return render(request, "depart_list.html", context)

def depart_add(request):
    """ add department """
    # 检查用户是否登陆，用户发来请求获取cookie随机字符串，拿来看看服务器session中有没有
    if request.method == "GET":
        return render(request, "depart_add.html")

    print(request.POST)
    title = request.POST.get("title")

    models.Department.objects.create(title=title)

    return redirect("/depart/list/")

def depart_delete(request):
    """ delete department """
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()

    return redirect("/depart/list/")

def depart_edit(request, nid):
    """ update department """
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id, row_object.title)
        return render(request, "depart_edit.html", {"row_object": row_object})
    else:
        title = request.POST.get("title")
        models.Department.objects.filter(id=nid).update(title=title)
        return redirect("/depart/list/")

def depart_multi(request):
    file_object = request.FILES.get("exc")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    for row_obj in sheet.iter_rows(min_row=2):
        if not row_obj[1].value:
            continue
        text = row_obj[1].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect("/depart/list/")


